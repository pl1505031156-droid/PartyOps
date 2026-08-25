"""深度审查报告中高价值问题的根因回归测试。"""

from __future__ import annotations

import asyncio
import hashlib
import io
import json
import logging
import os
import threading
import time
import uuid
from contextlib import contextmanager
from datetime import datetime, timedelta, timezone
from pathlib import Path
from types import SimpleNamespace

import pytest
from app import backups as backups_module
from app import main as main_module
from app import networking as networking_module
from app import notifications as notifications_module
from app import projections as projections_module
from app import recommendations as recommendations_module
from app import reports as reports_module
from app import scheduler
from app import workspace as workspace_module
from app.config import Settings, get_settings
from app.database import db_runtime
from app.enums import Sensitivity, TransferDirection, TransferStatus
from app.local_ai import LocalLlmRuntime
from app.login_throttle import LoginThrottle
from app.models import (
    AIModelPack,
    EventOutbox,
    LoginSession,
    Notification,
    PeriodReport,
    PeriodReportItem,
    ProjectionCheckpoint,
    SemanticIndexCheckpoint,
    Task,
    TaskParticipant,
    Transfer,
    TransferChunk,
    User,
    WorkspaceRoot,
)
from app.problems import ProblemException
from app.routers import admin as admin_router_module
from app.routers import fleet
from app.schemas import TaskUpdate
from app.task_service import update_task
from app.workspace import scan_root
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import event, func, select

from .conftest import create_task


def login_as(client: TestClient, username: str) -> None:
    response = client.post(
        "/api/v1/auth/login",
        json={"username": username, "password": "PartyOps@2026"},
    )
    assert response.status_code == 200, response.text


def test_workspace_scan_serializes_same_root(monkeypatch: pytest.MonkeyPatch) -> None:
    """自动扫描与立即同步不能同时写入同一个根目录索引。"""

    state = {"active": 0, "max_active": 0}
    state_lock = threading.Lock()
    start = threading.Barrier(3)
    completed: list[str] = []

    def fake_scan(_db, root):
        with state_lock:
            state["active"] += 1
            state["max_active"] = max(state["max_active"], state["active"])
        time.sleep(0.05)
        with state_lock:
            state["active"] -= 1
        return root.id

    monkeypatch.setattr(workspace_module, "_scan_root_locked", fake_scan)

    def run() -> None:
        start.wait(timeout=2)
        completed.append(
            workspace_module.scan_root(None, SimpleNamespace(id="same-root"))
        )

    workers = [threading.Thread(target=run, daemon=True) for _ in range(2)]
    for worker in workers:
        worker.start()
    start.wait(timeout=2)
    for worker in workers:
        worker.join(timeout=2)

    assert completed == ["same-root", "same-root"]
    assert state["max_active"] == 1


def test_collaborator_can_contribute_but_cannot_govern_task(
    client: TestClient,
    admin: dict,
    staff: dict,
) -> None:
    """协办人可贡献材料，但不能移交、删除、调整人员或流转状态。"""

    login_as(client, "admin")
    task = create_task(
        client,
        admin["id"],
        title="协办权限边界",
        collaborator_ids=[staff["id"]],
        materials=[{"category": "supporting", "name": "协办材料", "required": False}],
    )
    login_as(client, "staff")
    denied_update = client.patch(
        f"/api/v1/tasks/{task['id']}",
        headers={"If-Match": str(task["version"])},
        json={"owner_id": staff["id"]},
    )
    assert denied_update.status_code == 403
    assert denied_update.json()["code"] == "TASK_MANAGE_DENIED"
    denied_action = client.post(
        f"/api/v1/tasks/{task['id']}/actions",
        json={"action": "complete", "note": "越权完成"},
    )
    assert denied_action.status_code == 403
    assert denied_action.json()["code"] == "TASK_WORKFLOW_DENIED"
    denied_delete = client.delete(
        f"/api/v1/tasks/{task['id']}",
        headers={"If-Match": str(task["version"])},
    )
    assert denied_delete.status_code == 403

    contributed = client.post(
        f"/api/v1/tasks/{task['id']}/materials/{task['materials'][0]['id']}/versions",
        data={"stage": "draft", "is_final": "false", "note": "协办补充"},
        files={"file": ("协办说明.txt", io.BytesIO(b"collaboration"), "text/plain")},
    )
    assert contributed.status_code == 201, contributed.text
    login_as(client, "admin")


def test_final_material_and_archived_task_are_immutable(
    client: TestClient,
    admin: dict,
) -> None:
    """已有终稿不能被静默替换，归档快照形成后也不能继续上传。"""

    login_as(client, "admin")
    task = create_task(client, admin["id"], title="终稿锁定检查")
    material_id = task["materials"][0]["id"]
    final = client.post(
        f"/api/v1/tasks/{task['id']}/materials/{material_id}/versions",
        data={"stage": "submitted", "is_final": "true", "note": "正式终稿"},
        files={"file": ("终稿.txt", io.BytesIO(b"final-v1"), "text/plain")},
    )
    assert final.status_code == 201, final.text
    replacement = client.post(
        f"/api/v1/tasks/{task['id']}/materials/{material_id}/versions",
        data={"stage": "submitted", "is_final": "true", "note": "静默替换"},
        files={"file": ("终稿-2.txt", io.BytesIO(b"final-v2"), "text/plain")},
    )
    assert replacement.status_code == 409
    assert replacement.json()["code"] == "FINAL_VERSION_LOCKED"

    completed = client.post(
        f"/api/v1/tasks/{task['id']}/actions",
        json={"action": "complete", "note": "办理完成"},
    )
    assert completed.status_code == 200, completed.text
    archived = client.post(
        f"/api/v1/tasks/{task['id']}/actions",
        json={"action": "archive", "note": "形成归档快照"},
    )
    assert archived.status_code == 200, archived.text
    after_archive = client.post(
        f"/api/v1/tasks/{task['id']}/materials/{material_id}/versions",
        data={"stage": "draft", "is_final": "false", "note": "归档后补传"},
        files={"file": ("补传.txt", io.BytesIO(b"late"), "text/plain")},
    )
    assert after_archive.status_code == 409
    assert after_archive.json()["code"] == "TASK_ARCHIVED_IMMUTABLE"


def test_material_rollback_creates_audited_new_final_without_deleting_history(
    client: TestClient,
    admin: dict,
) -> None:
    """回退必须引用旧 Blob 新建版本，原终稿与全部过程版本继续可下载。"""

    login_as(client, "admin")
    task = create_task(client, admin["id"], title="材料回退检查")
    material_id = task["materials"][0]["id"]
    draft = client.post(
        f"/api/v1/tasks/{task['id']}/materials/{material_id}/versions",
        headers={"If-Match": str(task["version"])},
        data={"stage": "draft", "is_final": "false", "note": "首版"},
        files={"file": ("首版.txt", io.BytesIO(b"version-one"), "text/plain")},
    )
    assert draft.status_code == 201, draft.text
    task = draft.json()
    target = task["materials"][0]["versions"][0]
    final = client.post(
        f"/api/v1/tasks/{task['id']}/materials/{material_id}/versions",
        headers={"If-Match": str(task["version"])},
        data={"stage": "submitted", "is_final": "true", "note": "正式终稿"},
        files={"file": ("终稿.txt", io.BytesIO(b"version-two"), "text/plain")},
    )
    assert final.status_code == 201, final.text
    task = final.json()

    missing_reason = client.post(
        f"/api/v1/tasks/{task['id']}/materials/{material_id}/versions/{target['id']}/rollback",
        headers={"If-Match": str(task["version"])},
        json={"reason": ""},
    )
    assert missing_reason.status_code == 422
    stale = client.post(
        f"/api/v1/tasks/{task['id']}/materials/{material_id}/versions/{target['id']}/rollback",
        headers={"If-Match": str(task["version"] - 1)},
        json={"reason": "终稿内容有误，恢复首版"},
    )
    assert stale.status_code == 409
    assert stale.json()["code"] == "VERSION_CONFLICT"

    rolled_back = client.post(
        f"/api/v1/tasks/{task['id']}/materials/{material_id}/versions/{target['id']}/rollback",
        headers={"If-Match": str(task["version"])},
        json={"reason": "终稿内容有误，恢复首版"},
    )
    assert rolled_back.status_code == 200, rolled_back.text
    versions = rolled_back.json()["materials"][0]["versions"]
    assert [version["version_no"] for version in versions] == [3, 2, 1]
    assert [version["is_final"] for version in versions] == [True, False, False]
    assert versions[0]["note"] == "回退至 v1：终稿内容有误，恢复首版"
    assert (
        client.get(f"/api/v1/attachments/{versions[0]['id']}/download").content
        == b"version-one"
    )
    assert (
        client.get(f"/api/v1/attachments/{versions[1]['id']}/download").content
        == b"version-two"
    )
    already_final = client.post(
        f"/api/v1/tasks/{task['id']}/materials/{material_id}/versions/{versions[0]['id']}/rollback",
        headers={"If-Match": str(rolled_back.json()["version"])},
        json={"reason": "重复选择当前终稿"},
    )
    assert already_final.status_code == 409
    assert already_final.json()["code"] == "ATTACHMENT_ALREADY_FINAL"


def _completed_inbox_transfer(admin_id: str, content: bytes, name: str) -> Transfer:
    transfer = Transfer(
        direction="device_to_host",
        status="completed",
        original_name=name,
        size_bytes=len(content),
        sha256=hashlib.sha256(content).hexdigest(),
        requested_by=admin_id,
        expires_at=datetime.now(timezone.utc) + timedelta(days=7),
    )
    with db_runtime.session_factory() as db:
        db.add(transfer)
        db.commit()
        transfer_id = transfer.id
    get_settings().inbox_dir.joinpath(
        f"{transfer_id}-{fleet.safe_name(name)}"
    ).write_bytes(content)
    with db_runtime.session_factory() as db:
        return db.get(Transfer, transfer_id)


def test_inbox_attach_cannot_bypass_final_or_archive_lock(
    client: TestClient,
    admin: dict,
) -> None:
    """接收箱转材料必须与普通上传共用终稿和归档锁。"""

    login_as(client, "admin")
    task = create_task(client, admin["id"], title="接收箱终稿旁路检查")
    material_id = task["materials"][0]["id"]
    first = client.post(
        f"/api/v1/tasks/{task['id']}/materials/{material_id}/versions",
        data={"stage": "submitted", "is_final": "true", "note": "已锁定终稿"},
        files={"file": ("正式终稿.txt", io.BytesIO(b"official-final"), "text/plain")},
    )
    assert first.status_code == 201, first.text

    replacement = _completed_inbox_transfer(admin["id"], b"replacement", "替换终稿.txt")
    denied = client.post(
        f"/api/v1/transfers/{replacement.id}/attach",
        json={
            "target_type": "task_material",
            "target_id": material_id,
            "stage": "submitted",
            "is_final": True,
            "note": "不应绕过终稿锁",
        },
    )
    assert denied.status_code == 409, denied.text
    assert denied.json()["code"] == "FINAL_VERSION_LOCKED"

    assert (
        client.post(
            f"/api/v1/tasks/{task['id']}/actions",
            json={"action": "complete", "note": "办理完成"},
        ).status_code
        == 200
    )
    assert (
        client.post(
            f"/api/v1/tasks/{task['id']}/actions",
            json={"action": "archive", "note": "完成归档"},
        ).status_code
        == 200
    )
    late = _completed_inbox_transfer(admin["id"], b"late", "归档后补件.txt")
    archived = client.post(
        f"/api/v1/transfers/{late.id}/attach",
        json={
            "target_type": "task_material",
            "target_id": material_id,
            "stage": "draft",
            "is_final": False,
            "note": "不应写入归档事项",
        },
    )
    assert archived.status_code == 409, archived.text
    assert archived.json()["code"] == "TASK_ARCHIVED_IMMUTABLE"


def test_transfer_retention_reclaims_expired_part(admin: dict) -> None:
    """过期传输应释放临时文件、分块记录和配额。"""

    settings = get_settings()
    with db_runtime.session_factory() as db:
        transfer = Transfer(
            direction="host_to_device",
            status="queued",
            original_name="过期材料.zip",
            size_bytes=7,
            total_chunks=1,
            completed_chunks=1,
            requested_by=admin["id"],
            transit_path="expired.part",
            expires_at=datetime.now(timezone.utc) - timedelta(minutes=1),
        )
        db.add(transfer)
        db.flush()
        db.add(
            TransferChunk(
                transfer_id=transfer.id,
                chunk_no=0,
                sha256=hashlib.sha256(b"expired").hexdigest(),
                size_bytes=7,
            )
        )
        db.commit()
        part = settings.transfers_dir / f"{transfer.id}.part"
        part.write_bytes(b"expired")

        assert scheduler.cleanup_transfer_storage(db, settings) == 1
        db.commit()
        db.refresh(transfer)
        assert transfer.status == "expired"
        assert transfer.completed_chunks == 0
        assert (
            db.scalar(
                select(TransferChunk.id).where(TransferChunk.transfer_id == transfer.id)
            )
            is None
        )
    assert not part.exists()


def test_runtime_retention_bounds_database_and_disk_growth(
    tmp_path,
    admin: dict,
) -> None:
    """已读通知、过期会话、接收箱和临时导出必须有明确生命周期。"""

    now = datetime.now(timezone.utc)
    old = now - timedelta(days=10)
    runtime = SimpleNamespace(
        inbox_dir=tmp_path / "inbox",
        exports_dir=tmp_path / "exports",
        data_dir=tmp_path,
        notification_read_retention_days=5,
        session_retention_days=5,
        transient_record_retention_days=5,
        event_outbox_retention_days=5,
        inbox_handled_retention_days=5,
        inbox_unhandled_retention_days=5,
        export_retention_days=5,
        upgrade_backup_retention_days=5,
        upgrade_backup_keep=1,
    )
    runtime.inbox_dir.mkdir()
    runtime.exports_dir.mkdir()
    upgrade_root = runtime.data_dir / "upgrade-backups"
    upgrade_root.mkdir()

    with db_runtime.session_factory() as db:
        read_notice = Notification(
            user_id=admin["id"],
            notification_type="test",
            title="旧已读",
            dedupe_key=f"retention-read-{uuid.uuid4().hex}",
            read_at=old,
            created_at=old,
        )
        unread_notice = Notification(
            user_id=admin["id"],
            notification_type="test",
            title="旧未读",
            dedupe_key=f"retention-unread-{uuid.uuid4().hex}",
            created_at=old,
        )
        session = LoginSession(
            token_hash=hashlib.sha256(uuid.uuid4().bytes).hexdigest(),
            user_id=admin["id"],
            expires_at=old,
            last_seen_at=old,
            created_at=old,
        )
        handled = Transfer(
            direction="device_to_host",
            status="completed",
            original_name="已处理.txt",
            requested_by=admin["id"],
            expires_at=now + timedelta(days=1),
            handled_by=admin["id"],
            handled_at=old,
            created_at=old,
        )
        unhandled = Transfer(
            direction="device_to_host",
            status="completed",
            original_name="未处理.txt",
            requested_by=admin["id"],
            expires_at=now + timedelta(days=1),
            created_at=old,
        )
        db.add_all([read_notice, unread_notice, session, handled, unhandled])
        db.commit()
        ids = read_notice.id, unread_notice.id, session.id, handled.id, unhandled.id

        handled_path = runtime.inbox_dir / f"{handled.id}-已处理.txt"
        unhandled_path = runtime.inbox_dir / f"{unhandled.id}-未处理.txt"
        handled_path.write_bytes(b"handled")
        unhandled_path.write_bytes(b"unhandled")
        old_export = runtime.exports_dir / "old.zip"
        recent_export = runtime.exports_dir / "recent.zip"
        old_export.write_bytes(b"old")
        recent_export.write_bytes(b"recent")
        old_stamp = old.timestamp()
        old_export.touch()
        os.utime(old_export, (old_stamp, old_stamp))
        old_upgrade = upgrade_root / "old-run"
        new_upgrade = upgrade_root / "new-run"
        old_upgrade.mkdir()
        new_upgrade.mkdir()
        os.utime(old_upgrade, (old_stamp, old_stamp))

        scheduler.cleanup_runtime_retention(db, runtime, now=now)
        db.commit()

        assert db.get(Notification, ids[0]) is None
        assert db.get(Notification, ids[1]) is not None
        assert db.get(LoginSession, ids[2]) is None
        assert db.get(Transfer, ids[3]).error_code == "INBOX_RETAINED_COPY_EXPIRED"
        assert db.get(Transfer, ids[4]).error_code == "INBOX_EXPIRED"
    assert not handled_path.exists()
    assert not unhandled_path.exists()
    assert not old_export.exists()
    assert recent_export.exists()
    assert not old_upgrade.exists()
    assert new_upgrade.exists()


def test_projection_failure_uses_persisted_backoff(admin: dict, monkeypatch) -> None:
    """持续失败的投影不能每个调度周期重复冲击数据库。"""

    with db_runtime.session_factory() as db:
        event_item = EventOutbox(
            event_type="task.updated",
            entity_id=None,
            payload={"test": "projection-backoff"},
        )
        db.add(event_item)
        db.commit()
        checkpoint = db.get(ProjectionCheckpoint, projections_module.REPORT_PROJECTION)
        if checkpoint is None:
            checkpoint = ProjectionCheckpoint(name=projections_module.REPORT_PROJECTION)
            db.add(checkpoint)
        checkpoint.last_event_id = event_item.id - 1
        checkpoint.failed_count = 0
        checkpoint.status = "idle"
        checkpoint.last_run_at = None
        db.commit()

        calls = 0

        def broken_projection(*_args, **_kwargs):
            nonlocal calls
            calls += 1
            raise RuntimeError("deterministic projection failure")

        monkeypatch.setattr(projections_module, "_sync_reports", broken_projection)
        user = db.get(User, admin["id"])
        projections_module.process_report_projection(db, user)
        projections_module.process_report_projection(db, user)
        assert calls == 1


def test_overdue_notification_refreshes_single_row(
    client: TestClient,
    admin: dict,
    monkeypatch,
) -> None:
    """持续逾期只刷新一个未完成提醒，不得每天新增一行。"""

    task = create_task(
        client,
        admin["id"],
        title="通知有界增长检查",
        formal_due_at="2020-01-02T00:00:00Z",
        internal_due_at="2020-01-01T00:00:00Z",
    )
    moments = iter(
        [
            datetime(2026, 8, 8, tzinfo=timezone.utc),
            datetime(2026, 8, 9, tzinfo=timezone.utc),
        ]
    )
    monkeypatch.setattr(notifications_module, "utcnow", lambda: next(moments))
    with db_runtime.session_factory() as db:
        notifications_module.refresh_notifications(db)
        db.commit()
        notifications_module.refresh_notifications(db)
        db.commit()
        count = db.scalar(
            select(func.count(Notification.id)).where(
                Notification.user_id == admin["id"],
                Notification.entity_id == task["id"],
                Notification.notification_type == "overdue",
            )
        )
        assert count == 1


def test_invalid_legacy_quiet_hours_fall_back_safely() -> None:
    """升级前遗留的非法免打扰时间不能让通知设置接口变成 500。"""

    preference = SimpleNamespace(
        enabled=True,
        desktop_enabled=True,
        quiet_start="bad-value",
        quiet_end="99:77",
    )
    assert isinstance(
        notifications_module.desktop_notifications_allowed(
            preference,
            datetime(2026, 8, 9, 12, 0, tzinfo=timezone.utc),
        ),
        bool,
    )


def test_task_update_rechecks_version_inside_write_lock(
    client: TestClient,
    admin: dict,
) -> None:
    """两个会话同时编辑时，后进入写锁的旧版本不得覆盖新值。"""

    task = create_task(client, admin["id"], title="原子乐观锁检查")
    first_db = db_runtime.session_factory()
    stale_db = db_runtime.session_factory()
    try:
        first_task = first_db.get(Task, task["id"])
        stale_task = stale_db.get(Task, task["id"])
        first_user = first_db.get(User, admin["id"])
        stale_user = stale_db.get(User, admin["id"])
        version = first_task.version
        update_task(
            first_db,
            first_task,
            TaskUpdate(title="先提交的新标题"),
            version,
            first_user,
        )
        with pytest.raises(ProblemException) as captured:
            update_task(
                stale_db,
                stale_task,
                TaskUpdate(title="不应覆盖的新标题"),
                version,
                stale_user,
            )
        assert captured.value.code == "VERSION_CONFLICT"
    finally:
        first_db.close()
        stale_db.close()
    with db_runtime.session_factory() as db:
        assert db.get(Task, task["id"]).title == "先提交的新标题"


def test_batch_transition_is_atomic_and_owner_participant_stays_consistent(
    client: TestClient,
    admin: dict,
    staff: dict,
) -> None:
    """批量中任一事项失败应全部回滚，转交后 OWNER 参与关系必须同步。"""

    valid = create_task(client, admin["id"], title="批量原子-可变更")
    invalid = create_task(client, admin["id"], title="批量原子-应阻断")
    completed = client.post(
        f"/api/v1/tasks/{invalid['id']}/actions",
        json={"action": "complete", "note": "先完成，制造不兼容状态"},
    )
    assert completed.status_code == 200, completed.text

    failed = client.post(
        "/api/v1/tasks/batch",
        json={
            "task_ids": [valid["id"], invalid["id"]],
            "status": "waiting_feedback",
            "note": "验证整批回滚",
        },
    )
    assert failed.status_code == 409, failed.text
    assert client.get(f"/api/v1/tasks/{valid['id']}").json()["status"] == "in_progress"

    reassigned = client.post(
        "/api/v1/tasks/batch",
        json={"task_ids": [invalid["id"]], "owner_id": staff["id"], "note": "批量转交"},
    )
    assert reassigned.status_code == 200, reassigned.text
    with db_runtime.session_factory() as db:
        owner_ids = set(
            db.scalars(
                select(TaskParticipant.user_id).where(
                    TaskParticipant.task_id == invalid["id"],
                    TaskParticipant.role == "owner",
                )
            ).all()
        )
        assert db.get(Task, invalid["id"]).owner_id == staff["id"]
        assert owner_ids == {staff["id"]}


def test_archive_record_survives_removed_template_field(
    client: TestClient,
) -> None:
    """管理员删去模板字段后，旧档案仍可修订且保留历史字段值。"""

    suffix = uuid.uuid4().hex[:8]
    category = client.post(
        "/api/v1/archives/categories",
        json={
            "name": f"模板演进-{suffix}",
            "code": f"template_evolution_{suffix}",
            "record_mode": "document",
            "access_mode": "all_users",
            "field_schema": [
                {
                    "key": "legacy_field",
                    "label": "历史字段",
                    "type": "text",
                    "required": False,
                }
            ],
        },
    )
    assert category.status_code == 201, category.text
    record = client.post(
        "/api/v1/archives/records",
        json={
            "category_id": category.json()["id"],
            "archive_year": 2026,
            "title": "包含历史字段的档案",
            "custom_fields": {"legacy_field": "不得丢失"},
        },
    )
    assert record.status_code == 201, record.text
    changed = client.patch(
        f"/api/v1/archives/categories/{category.json()['id']}",
        headers={"If-Match": str(category.json()["version"])},
        json={"field_schema": []},
    )
    assert changed.status_code == 200, changed.text

    revised = client.patch(
        f"/api/v1/archives/records/{record.json()['id']}",
        headers={"If-Match": str(record.json()["version"])},
        json={"summary": "模板变化后仍可修订", "change_note": "兼容历史字段"},
    )
    assert revised.status_code == 200, revised.text
    assert revised.json()["custom_fields"]["legacy_field"] == "不得丢失"


def test_backup_hashing_does_not_hold_global_write_lock(
    admin: dict,
    monkeypatch,
) -> None:
    """大附件哈希和压缩期间，普通业务写锁仍应可及时获取。"""

    settings = get_settings()
    attachment = settings.attachments_dir / f"lock-scope-{uuid.uuid4().hex}.bin"
    attachment.write_bytes(b"backup-lock-scope")
    hashing = threading.Event()
    failures: list[BaseException] = []
    original_sha = backups_module.sha256_file

    def slow_attachment_hash(path):
        if Path(path) == attachment:
            hashing.set()
            time.sleep(0.2)
        return original_sha(path)

    def run_backup() -> None:
        try:
            with db_runtime.session_factory() as db:
                actor = db.get(User, admin["id"])
                backups_module.create_backup(db, actor, kind="manual")
        except BaseException as exc:  # pragma: no cover - 线程故障回传
            failures.append(exc)

    monkeypatch.setattr(backups_module, "sha256_file", slow_attachment_hash)
    worker = threading.Thread(target=run_backup, daemon=True)
    worker.start()
    assert hashing.wait(timeout=3)
    started = time.perf_counter()
    with db_runtime.write_lock:
        elapsed = time.perf_counter() - started
    worker.join(timeout=10)
    attachment.unlink(missing_ok=True)
    assert not worker.is_alive()
    assert failures == []
    assert elapsed < 0.08


def test_structured_log_formatter_escapes_untrusted_text() -> None:
    """日志消息中的引号和换行必须仍形成单行合法 JSON。"""

    record = logging.LogRecord(
        "partyops.test",
        logging.WARNING,
        __file__,
        1,
        '用户输入"伪造"\n第二行',
        (),
        None,
    )
    encoded = main_module.JsonLogFormatter().format(record)
    assert "\n" not in encoded
    payload = json.loads(encoded)
    assert payload["message"] == '用户输入"伪造"\n第二行'


def test_invalid_trace_id_is_replaced(client: TestClient) -> None:
    """客户端不能把任意文本带入响应、日志和诊断关联号。"""

    response = client.get(
        "/api/v1/health",
        headers={"X-Trace-Id": 'forged"trace'},
    )
    assert response.status_code == 200
    assert (
        str(uuid.UUID(response.headers["X-Trace-Id"])) == response.headers["X-Trace-Id"]
    )


def test_production_lan_rejects_plain_http() -> None:
    """生产局域网绑定必须启用 TLS，本机环回冒烟仍可显式使用 HTTP。"""

    with pytest.raises(RuntimeError):
        networking_module.validate_transport_security(
            host="192.168.10.20",
            production=True,
            tls_enabled=False,
        )
    networking_module.validate_transport_security(
        host="127.0.0.1",
        production=True,
        tls_enabled=False,
    )


def test_system_status_reports_low_disk_as_degraded(
    client: TestClient,
    admin: dict,
    monkeypatch,
) -> None:
    """磁盘不足时诊断页不能继续显示 status=ok。"""

    login_as(client, "admin")
    monkeypatch.setattr(
        admin_router_module.shutil,
        "disk_usage",
        lambda _path: SimpleNamespace(
            total=10 * 1024**3, used=9 * 1024**3, free=512 * 1024**2
        ),
    )
    response = client.get("/api/v1/admin/system-status")
    assert response.status_code == 200, response.text
    assert response.json()["ready"] is False
    assert response.json()["status"] == "degraded"
    assert response.json()["readiness"]["storage_headroom"] is False


def test_local_llm_start_failure_uses_backoff(monkeypatch) -> None:
    """本地运行时缺失后不得在每次请求中重复走完整启动流程。"""

    runtime = LocalLlmRuntime()
    pack = SimpleNamespace(id="backoff-pack", model_id="qwen-test")
    monkeypatch.setattr(LocalLlmRuntime, "_binary", staticmethod(lambda: None))
    with pytest.raises(ProblemException) as first:
        runtime._ensure_started(pack)
    assert first.value.code == "LOCAL_LLM_RUNTIME_MISSING"
    with pytest.raises(ProblemException) as second:
        runtime._ensure_started(pack)
    assert second.value.code == "LOCAL_LLM_START_BACKOFF"
    assert int(second.value.headers["Retry-After"]) >= 1


def test_login_throttle_flood_does_not_evict_locked_account(tmp_path) -> None:
    """随机用户名灌入只能淘汰低价值噪声，不能挤掉真实账号锁定。"""

    settings = Settings(
        data_dir=tmp_path,
        login_account_failure_limit=3,
        login_ip_failure_limit=200,
        login_throttle_max_entries=128,
        login_window_seconds=600,
        login_lock_seconds=600,
    )
    throttle = LoginThrottle()
    for current in (1.0, 2.0, 3.0):
        throttle.record_failure(
            "protected-admin",
            "192.0.2.10",
            now=current,
            settings=settings,
        )
    for index in range(300):
        throttle.record_failure(
            f"noise-{index}",
            f"198.51.100.{index}",
            now=4.0 + index / 1000,
            settings=settings,
        )
    assert (
        throttle.retry_after(
            "protected-admin",
            "192.0.2.10",
            now=5.0,
            settings=settings,
        )
        > 0
    )
    assert len(throttle._states) <= settings.login_throttle_max_entries


def test_exclusive_maintenance_waits_and_rejects_new_sessions(admin: dict) -> None:
    """恢复闸门必须等待在途事务，并在换库窗口拒绝新数据库访问。"""

    active_db = db_runtime.session_factory()
    assert (
        active_db.scalar(select(User.id).where(User.id == admin["id"])) == admin["id"]
    )
    entered = threading.Event()
    release = threading.Event()
    failures: list[BaseException] = []

    def run_maintenance() -> None:
        try:
            with db_runtime.exclusive_maintenance(timeout_seconds=2):
                entered.set()
                release.wait(timeout=2)
        except BaseException as exc:  # pragma: no cover - 仅用于回传线程故障
            failures.append(exc)

    worker = threading.Thread(target=run_maintenance, daemon=True)
    worker.start()
    assert not entered.wait(timeout=0.05)
    active_db.close()
    assert entered.wait(timeout=1)

    with db_runtime.session_factory() as blocked_db:
        with pytest.raises(ProblemException) as captured:
            blocked_db.scalar(select(User.id).limit(1))
    assert captured.value.code == "DATABASE_MAINTENANCE"

    release.set()
    worker.join(timeout=2)
    assert not worker.is_alive()
    assert failures == []
    with db_runtime.session_factory() as healthy_db:
        assert healthy_db.scalar(select(User.id).where(User.id == admin["id"]))


def test_heartbeat_keeps_enrollment_recovery(client: TestClient, admin: dict) -> None:
    """首次响应丢失后，终端应能在心跳后继续幂等恢复入网结果。"""

    device_name = f"审查恢复电脑-{uuid.uuid4().hex[:8]}"
    enrollment = client.post(
        "/api/v1/admin/devices/enrollments",
        json={"name": device_name},
    )
    assert enrollment.status_code == 201, enrollment.text
    payload = {
        "code": enrollment.json()["code"],
        "name": device_name,
        "architecture": "amd64",
        "platform": "windows",
        "kernel": "10.0.26100",
        "agent_version": "1.4.3",
        "local_username": "review-user",
        "disk_free_bytes": 10_000_000_000,
    }
    enrolled = client.post("/api/v1/devices/enroll", json=payload)
    assert enrolled.status_code == 201, enrolled.text

    heartbeat = client.post(
        "/api/v1/devices/heartbeat",
        headers={"X-PartyOps-Device-Token": enrolled.json()["device_token"]},
        json={
            "architecture": "amd64",
            "platform": "windows",
            "kernel": "10.0.26100",
            "agent_version": "1.4.3",
            "local_username": "review-user",
            "disk_free_bytes": 9_000_000_000,
        },
    )
    assert heartbeat.status_code == 200, heartbeat.text

    recovered = client.post("/api/v1/devices/enroll", json=payload)
    assert recovered.status_code == 201, recovered.text
    assert recovered.json()["device_id"] == enrolled.json()["device_id"]
    assert recovered.json()["device_token"] == enrolled.json()["device_token"]


@pytest.mark.asyncio
async def test_scheduler_cycle_does_not_block_event_loop(monkeypatch) -> None:
    """同步调度任务必须在线程中运行，不能冻结 API 所在事件循环。"""

    class EmptyScalars:
        def all(self) -> list:
            return []

    class FakeSession:
        def scalar(self, _statement):
            return None

        def scalars(self, _statement) -> EmptyScalars:
            return EmptyScalars()

        def execute(self, _statement) -> None:
            return None

        def commit(self) -> None:
            return None

    @contextmanager
    def fake_session_factory():
        yield FakeSession()

    class OneCycleStop:
        def __init__(self) -> None:
            self.stopped = False

        def is_set(self) -> bool:
            return self.stopped

        async def wait(self) -> bool:
            self.stopped = True
            return True

    monkeypatch.setattr(scheduler.db_runtime, "session_factory", fake_session_factory)
    monkeypatch.setattr(
        scheduler, "refresh_notifications", lambda _db: time.sleep(0.15)
    )

    started = time.perf_counter()
    task = asyncio.create_task(scheduler.scheduler_loop(OneCycleStop()))
    await asyncio.sleep(0.02)
    responsiveness = time.perf_counter() - started
    await task

    assert responsiveness < 0.08


@pytest.mark.asyncio
async def test_upload_final_hash_does_not_block_event_loop(
    tmp_path, monkeypatch
) -> None:
    """末块整文件校验和落盘必须离开事件循环。"""

    content = b"deep-review"
    digest = hashlib.sha256(content).hexdigest()
    settings = SimpleNamespace(
        transfers_dir=tmp_path / "transfers",
        inbox_dir=tmp_path / "inbox",
        transfer_max_file_gb=20,
    )
    settings.transfers_dir.mkdir()
    settings.inbox_dir.mkdir()
    transfer = SimpleNamespace(
        id="deep-review-transfer",
        status="queued",
        source_device_id="review-device",
        destination_device_id=None,
        expires_at=datetime.now(timezone.utc) + timedelta(hours=1),
        total_chunks=1,
        chunk_size=len(content),
        size_bytes=len(content),
        completed_chunks=0,
        sha256=digest,
        direction="device_to_host",
        original_name="审查材料.txt",
        transit_path="",
        result_sha256="",
        version=1,
        error_code="",
        error_message="",
    )

    class FakeDatabase:
        def __init__(self) -> None:
            self.chunk = None

        def get(self, _model, _identity):
            return transfer

        def scalar(self, _statement):
            return 1 if self.chunk is not None else None

        def add(self, value) -> None:
            self.chunk = value

        def flush(self) -> None:
            return None

        def commit(self) -> None:
            return None

    class FakeRequest:
        async def body(self) -> bytes:
            return content

    original_sha256_path = fleet.sha256_path

    def slow_sha256_path(path):
        time.sleep(0.15)
        return original_sha256_path(path)

    monkeypatch.setattr(
        fleet,
        "authenticated_device",
        lambda _token, _db: SimpleNamespace(id="review-device"),
    )
    monkeypatch.setattr(
        fleet, "transfer_sources_still_allowed", lambda _db, _transfer: True
    )
    monkeypatch.setattr(fleet, "ensure_transfer_storage_available", lambda _size: None)
    monkeypatch.setattr(fleet, "get_settings", lambda: settings)
    monkeypatch.setattr(fleet, "sha256_path", slow_sha256_path)

    started = time.perf_counter()
    task = asyncio.create_task(
        fleet.upload_chunk(
            transfer.id,
            0,
            FakeRequest(),
            token="device-token",
            chunk_sha256=digest,
            db=FakeDatabase(),
        )
    )
    await asyncio.sleep(0.02)
    responsiveness = time.perf_counter() - started
    result = await task

    assert responsiveness < 0.08
    assert result["status"] == "transferring"


def test_workspace_scan_commits_each_batch(tmp_path, admin: dict) -> None:
    """大型目录扫描必须分批释放 SQLite 写锁。"""

    root_path = tmp_path / "large-root"
    root_path.mkdir()
    for index in range(501):
        (root_path / f"document-{index:04d}.txt").touch()

    with db_runtime.session_factory() as db:
        root = WorkspaceRoot(
            name=f"批量提交-{uuid.uuid4().hex[:8]}",
            absolute_path=str(root_path),
            selection_mode="all",
            included_paths=["."],
            created_by=admin["id"],
        )
        db.add(root)
        db.commit()
        commits = 0

        def record_commit(_connection) -> None:
            nonlocal commits
            commits += 1

        event.listen(db.bind, "commit", record_commit)
        try:
            result = scan_root(db, root)
        finally:
            event.remove(db.bind, "commit", record_commit)

    assert result.files == 501
    assert commits >= 3


def test_recurrence_exception_matches_same_instant_with_timezone_offset(
    client: TestClient,
    admin: dict,
) -> None:
    """同一 UTC 时刻使用不同时区写入时，跳过规则仍应准确命中。"""

    suffix = uuid.uuid4().hex[:8]
    template = client.post(
        "/api/v1/templates",
        json={
            "name": f"时区模板-{suffix}",
            "category": "专项",
            "task_type": "standard",
            "description": "验证周期例外的时区归一化。",
            "steps": ["核对规则"],
            "materials": [],
        },
    )
    assert template.status_code == 201, template.text
    rule = client.post(
        "/api/v1/recurrences",
        json={
            "name": f"时区规则-{suffix}",
            "template_id": template.json()["id"],
            "owner_id": admin["id"],
            "kind": "yearly",
            "internal_lead_days": 0,
            "next_run_at": "2020-01-01T00:00:00Z",
            "end_at": "2020-01-01T00:00:00Z",
        },
    )
    assert rule.status_code == 201, rule.text
    exception = client.post(
        f"/api/v1/recurrences/{rule.json()['id']}/exceptions",
        headers={"If-Match": str(rule.json()["version"])},
        json={
            "occurrence_at": "2020-01-01T08:00:00+08:00",
            "action": "skip",
            "reason": "同一时刻的跨时区跳过验证",
        },
    )
    assert exception.status_code == 201, exception.text

    generated = client.post("/api/v1/recurrences/run-due")
    assert generated.status_code == 200, generated.text
    with db_runtime.session_factory() as db:
        task_id = db.scalar(
            select(Task.id).where(Task.recurrence_rule_id == rule.json()["id"])
        )
    assert task_id is None


def test_semantic_candidates_do_not_starve_older_unindexed_tasks(
    client: TestClient,
    admin: dict,
) -> None:
    """最新对象已有检查点时，旧的未索引对象仍必须进入增量队列。"""

    older = create_task(client, admin["id"], title=f"待补旧索引-{uuid.uuid4().hex[:8]}")
    create_task(client, admin["id"], title=f"已索引新事项-{uuid.uuid4().hex[:8]}")
    with db_runtime.session_factory() as db:
        pack = AIModelPack(
            name="语义公平性测试包",
            version="1",
            model_id=f"test-{uuid.uuid4().hex}",
            filename=f"test-{uuid.uuid4().hex}.partyops-modelpack",
            install_key=f"test-{uuid.uuid4().hex}",
            sha256="0" * 64,
            capabilities=["embedding"],
            signature_valid=True,
            created_by=admin["id"],
        )
        db.add(pack)
        db.flush()
        tasks = list(
            db.scalars(
                select(Task).where(
                    Task.deleted_at.is_(None),
                    Task.sensitivity == Sensitivity.NORMAL,
                )
            ).all()
        )
        for task in tasks:
            if task.id == older["id"]:
                continue
            db.add(
                SemanticIndexCheckpoint(
                    object_type="task",
                    object_id=task.id,
                    object_version=task.version,
                    model_pack_id=pack.id,
                    embedding_blob=b"indexed",
                    content_sha256="1" * 64,
                )
            )
        db.flush()

        candidates = recommendations_module._index_candidates(db, 25, pack.id)
        assert any(item[0] == "task" and item[1] == older["id"] for item in candidates)
        db.rollback()


def test_semantic_batch_isolates_one_bad_item_without_losing_good_item(
    monkeypatch,
    admin: dict,
) -> None:
    """批量推理失败应二分隔离坏条目，正常条目仍建立有效向量。"""

    with db_runtime.session_factory() as db:
        pack = AIModelPack(
            name="语义故障隔离测试包",
            version="1",
            model_id=f"test-{uuid.uuid4().hex}",
            filename=f"test-{uuid.uuid4().hex}.partyops-modelpack",
            install_key=f"test-{uuid.uuid4().hex}",
            sha256="0" * 64,
            capabilities=["embedding"],
            signature_valid=True,
            created_by=admin["id"],
        )
        db.add(pack)
        db.flush()
        monkeypatch.setattr(
            recommendations_module,
            "local_ai_readiness",
            lambda _db, capability: {"ready": capability == "embedding"},
        )
        monkeypatch.setattr(
            recommendations_module,
            "active_model_pack",
            lambda _db, _capability: pack,
        )
        monkeypatch.setattr(
            recommendations_module,
            "_index_candidates",
            lambda _db, _limit, _pack_id: [
                ("task", "good-item", 1, "正常文本"),
                ("task", "bad-item", 1, "坏文本"),
            ],
        )

        def encode(_pack, texts, **_kwargs):
            if "坏文本" in texts:
                raise ProblemException(503, "EMBEDDING_BAD_ITEM", "测试坏条目")
            return [b"good-vector" for _text in texts]

        monkeypatch.setattr(recommendations_module.embedding_runtime, "encode", encode)
        indexed = recommendations_module.index_semantic_batch(db, limit=2)
        checkpoints = {
            item.object_id: item
            for item in db.scalars(
                select(SemanticIndexCheckpoint).where(
                    SemanticIndexCheckpoint.model_pack_id == pack.id
                )
            ).all()
        }

        assert indexed == 1
        assert checkpoints["good-item"].embedding_blob == b"good-vector"
        assert checkpoints["bad-item"].embedding_blob is None
        db.rollback()


def test_published_report_is_immutable_and_exports_its_snapshot(
    client: TestClient,
    admin: dict,
) -> None:
    """发布即形成不可变业务版本；旧数据漂移时导出仍以发布快照为准。"""

    login_as(client, "admin")
    anchor_year = 2100 + int(uuid.uuid4().hex[:4], 16) % 6800
    created = client.post(
        "/api/v1/period-reports",
        json={
            "period_type": "week",
            "anchor_at": f"{anchor_year:04d}-01-05T00:00:00Z",
            "title": f"发布锁定回归-{uuid.uuid4().hex[:8]}",
            "summary": "发布时说明",
            "auto_fill": False,
        },
    )
    assert created.status_code == 201, created.text
    report = created.json()
    added = client.post(
        f"/api/v1/period-reports/{report['id']}/items",
        headers={"If-Match": str(report["version"])},
        json={
            "section": "completed",
            "source_type": "manual",
            "title": "发布时条目",
            "content": "发布时内容",
            "sort_order": 1,
        },
    )
    assert added.status_code == 201, added.text
    current = client.get(f"/api/v1/period-reports/{report['id']}").json()
    published = client.post(
        f"/api/v1/period-reports/{report['id']}/actions",
        headers={"If-Match": str(current["version"])},
        json={"action": "publish", "note": "形成发布快照"},
    )
    assert published.status_code == 200, published.text

    denied_report = client.patch(
        f"/api/v1/period-reports/{report['id']}",
        headers={"If-Match": str(published.json()["version"])},
        json={"summary": "发布后漂移"},
    )
    assert denied_report.status_code == 409
    assert denied_report.json()["code"] == "REPORT_PUBLISHED"
    denied_item = client.patch(
        f"/api/v1/period-reports/{report['id']}/items/{added.json()['id']}",
        headers={"If-Match": str(added.json()["version"])},
        json={"title": "发布后漂移"},
    )
    assert denied_item.status_code == 409
    assert denied_item.json()["code"] == "REPORT_PUBLISHED"

    with db_runtime.session_factory() as db:
        stored = db.get(PeriodReport, report["id"])
        stored_item = db.get(PeriodReportItem, added.json()["id"])
        assert stored is not None and stored_item is not None
        stored_item.title = "模拟旧版本实时表漂移"
        db.flush()
        exported = reports_module.export_period_xlsx(db, stored)
        workbook = load_workbook(exported, read_only=True)
        values = [
            row[1] for row in workbook.active.iter_rows(min_row=2, values_only=True)
        ]
        workbook.close()
        exported.unlink(missing_ok=True)
        assert "发布时条目" in values
        assert "模拟旧版本实时表漂移" not in values
        db.rollback()


def test_workspace_root_delete_rejects_active_transfer_and_preserves_history(
    client: TestClient,
    admin: dict,
    tmp_path: Path,
) -> None:
    """移除共享根不得以外键 500 失败，也不能破坏活动传输或历史记录。"""

    login_as(client, "admin")
    root_path = tmp_path / "待删除活动传输目录"
    root_path.mkdir()
    created = client.post(
        "/api/v1/workspace/roots",
        json={
            "name": f"活动传输-{uuid.uuid4().hex[:8]}",
            "absolute_path": str(root_path),
        },
    )
    assert created.status_code == 201, created.text
    root = created.json()
    with db_runtime.session_factory() as db:
        transfer = Transfer(
            direction=TransferDirection.HOST_TO_DEVICE,
            status=TransferStatus.QUEUED,
            destination_root_id=root["id"],
            original_name="待交付材料.zip",
            requested_by=admin["id"],
            expires_at=datetime.now(timezone.utc) + timedelta(days=1),
        )
        db.add(transfer)
        db.commit()
        transfer_id = transfer.id

    impact = client.get(f"/api/v1/workspace/roots/{root['id']}/deletion-impact")
    assert impact.status_code == 200, impact.text
    assert impact.json()["active_transfers"] == 1
    assert impact.json()["physical_delete"] is False

    denied = client.request(
        "DELETE",
        f"/api/v1/workspace/roots/{root['id']}",
        headers={"If-Match": str(root["version"])},
        json={"reason": "目录改由新的共享位置承接"},
    )
    assert denied.status_code == 409
    assert denied.json()["code"] == "ROOT_HAS_ACTIVE_TRANSFERS"

    with db_runtime.session_factory() as db:
        transfer = db.get(Transfer, transfer_id)
        assert transfer is not None
        transfer.status = TransferStatus.COMPLETED
        db.commit()
    removed = client.request(
        "DELETE",
        f"/api/v1/workspace/roots/{root['id']}",
        headers={"If-Match": str(root["version"])},
        json={"reason": "目录改由新的共享位置承接"},
    )
    assert removed.status_code == 200, removed.text
    with db_runtime.session_factory() as db:
        historical = db.get(Transfer, transfer_id)
        assert historical is not None
        assert historical.destination_root_id == root["id"]
        preserved_root = db.get(WorkspaceRoot, root["id"])
        assert preserved_root is not None
        assert preserved_root.enabled is False
    disabled = client.get("/api/v1/workspace/roots", params={"lifecycle": "disabled"})
    assert disabled.status_code == 200, disabled.text
    assert root["id"] in [item["id"] for item in disabled.json()]

    restored = client.post(
        f"/api/v1/workspace/roots/{root['id']}/restore",
        headers={"If-Match": str(root["version"] + 1)},
        json={"reason": "管理员核对路径后恢复目录"},
    )
    assert restored.status_code == 200, restored.text
    assert restored.json()["root_id"] == root["id"]
    assert restored.json()["scan_required"] is True
    with db_runtime.session_factory() as db:
        assert db.get(WorkspaceRoot, root["id"]).enabled is True


def test_global_search_pushes_task_limit_into_sql(
    client: TestClient,
    admin: dict,
) -> None:
    """全局搜索不得先加载整张事项表再在 Python 中截断。"""

    login_as(client, "admin")
    create_task(client, admin["id"], title=f"检索下推-{uuid.uuid4().hex[:8]}")
    statements: list[str] = []

    def capture(_connection, _cursor, statement, _parameters, _context, _many) -> None:
        normalized = " ".join(statement.upper().split())
        if normalized.startswith("SELECT") and " FROM TASKS " in normalized:
            statements.append(normalized)

    with db_runtime.session_factory() as db:
        bind = db.get_bind()
    event.listen(bind, "before_cursor_execute", capture)
    try:
        response = client.get(
            "/api/v1/global-search", params={"q": "检索下推", "limit": 5}
        )
    finally:
        event.remove(bind, "before_cursor_execute", capture)

    assert response.status_code == 200, response.text
    assert statements
    assert " LIMIT " in statements[0]
